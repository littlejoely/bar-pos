"""
订单管理路由
"""
from flask import Blueprint, jsonify, request, send_file
import json
import os
import tempfile
import re
import smtplib
from io import BytesIO
from email.message import EmailMessage
from datetime import datetime
from routes.voucher import load_vouchers
from routes.system_settings import production_ticket_enabled

order_bp = Blueprint('order', __name__)

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
ORDERS_FILE = os.path.join(DATA_DIR, 'orders.json')
TABLES_FILE = os.path.join(DATA_DIR, 'tables.json')


def atomic_json_save(path, data):
    fd, temp_path = tempfile.mkstemp(prefix='.pos-bar-', suffix='.json', dir=DATA_DIR)
    try:
        with os.fdopen(fd, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(temp_path, path)
    except Exception:
        if os.path.exists(temp_path):
            os.unlink(temp_path)
        raise

def load_orders():
    try:
        with open(ORDERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)['orders']
    except:
        return []

def save_orders(orders):
    atomic_json_save(ORDERS_FILE, {'orders': orders})

def load_tables():
    with open(TABLES_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)['tables']

def save_tables(tables):
    with open(TABLES_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    data['tables'] = tables
    atomic_json_save(TABLES_FILE, data)

def update_table_status(table_id, status, reset=False, order_id=None):
    tables = load_tables()
    table = next((t for t in tables if t['id'] == table_id), None)
    if not table:
        return

    table['status'] = status
    if reset:
        table['guests'] = 0
        table['opened_at'] = None
        table['order_id'] = None
    elif order_id is not None:
        table['order_id'] = order_id

    save_tables(tables)

def create_pending_order(table_id, orders):
    tables = load_tables()
    table = next((t for t in tables if t['id'] == table_id), None)
    order_id = generate_order_id(table_id)
    table_guests = table.get('guests', 0) if table else 0
    if table_guests == 0 and table:
        table_guests = table.get('default_guests', 0) or 0
        if table_guests:
            table['guests'] = table_guests
            save_tables(tables)
    order = {
        'id': order_id,
        'table_id': table_id,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'items': [],
        'total': 0,
        'guests': table_guests,
        'status': 'pending'
    }
    orders.append(order)
    update_table_status(table_id, 'occupied', order_id=order_id)
    return order

def find_active_order(orders, table_id):
    return next((
        o for o in orders
        if o['table_id'] == table_id and o.get('status') in ['pending', 'submitted', 'served']
    ), None)


def next_order_line_id(order):
    """为同一菜单商品的多轮加菜生成独立订单行 ID。"""
    next_id = max(1000000, int(order.get('next_line_id', 1000000) or 1000000))
    existing_ids = {item.get('id') for item in order.get('items', [])}
    while next_id in existing_ids:
        next_id += 1
    order['next_line_id'] = next_id + 1
    return next_id


def item_line_total(item):
    if item.get('returned') is True:
        return 0
    qty = item.get('quantity', 0)
    price = item['price']
    gift_qty = item.get('gift_quantity', 0) or 0
    if gift_qty > qty:
        gift_qty = qty
    chargeable_qty = qty - gift_qty
    line = price * chargeable_qty
    discount = item.get('discount', 0) or 0
    reduction = item.get('reduction', 0) or 0
    if discount > 0:
        line = line * (1 - discount / 100)
    if reduction > 0:
        line = line - reduction
    return max(0, line)


def recalc_order(order):
    """重算每项 line_total 和订单 total，考虑折扣与减免。"""
    for item in order.get('items', []):
        item['line_total'] = round(item_line_total(item), 2)
    order['total'] = round(sum(item_line_total(i) for i in order.get('items', [])), 2)


def committed_order_total(order):
    return round(sum(
        item_line_total(item)
        for item in order.get('items', [])
        if item.get('addition_pending') is not True
    ), 2)


ALLOWED_PAYMENT_METHODS = ['微信支付', '支付宝', '现金', '银行卡', '挂账', '其他']


def voucher_financials(voucher):
    """拆分优惠券面值、实际收入和优惠金额。"""
    voucher = voucher or {}
    face_amount = round(voucher.get('amount', 0) or 0, 2)
    items = voucher.get('items') or []
    if not items:
        income = round(voucher.get('income_amount', face_amount) or 0, 2)
        income = min(max(0, income), face_amount)
        return face_amount, income, round(face_amount - income, 2)

    try:
        definitions = {v.get('id'): v for v in load_vouchers()}
    except Exception:
        definitions = {}
    income = 0
    for item in items:
        quantity = item.get('quantity', 0) or 0
        definition = definitions.get(item.get('id'), {})
        sale_price = item.get('sale_price')
        if sale_price is None:
            sale_price = definition.get('sale_price', item.get('face_value', 0))
            item['sale_price'] = round(float(sale_price or 0), 2)
        income += (sale_price or 0) * quantity
    income = round(min(max(0, income), face_amount), 2)
    return face_amount, income, round(face_amount - income, 2)


def compute_checkout_summary(order):
    """计算整单层面的应收/已收/未付。

    返回字段:
      subtotal          菜品小计 (= order.total)
      order_discount    整单折扣百分比
      discount_amount   整单折扣金额
      order_reduction   整单减免金额
      voucher_name      抵扣券名称
      voucher_amount    抵扣券面值
      voucher_income_amount   优惠券实际收入
      voucher_discount_amount 优惠券优惠金额
      round_down        抹零金额
      payable           应收金额 (抹零后)
      payments          已收款记录列表
      paid_total        已收总额
      balance_due       未付金额 (负数表示多收)
    """
    subtotal = round(order.get('total', 0) or 0, 2)
    order_discount = order.get('order_discount', 0) or 0
    if order_discount < 0 or order_discount > 100:
        order_discount = 0
    discount_amount = round(subtotal * order_discount / 100, 2)
    after_discount = round(subtotal - discount_amount, 2)

    order_reduction = round(order.get('order_reduction', 0) or 0, 2)
    if order_reduction < 0:
        order_reduction = 0
    after_reduction = round(after_discount - order_reduction, 2)

    voucher = order.get('voucher') or {}
    voucher_name = voucher.get('name') or ''
    voucher_amount, voucher_income_amount, voucher_discount_amount = voucher_financials(voucher)
    voucher_items = voucher.get('items') or []
    if voucher_amount < 0:
        voucher_amount = 0
    after_voucher_discount = round(after_reduction - voucher_discount_amount, 2)

    round_down = round(order.get('round_down', 0) or 0, 2)
    if round_down < 0:
        round_down = 0
    payable = round(after_voucher_discount - round_down, 2)
    if payable < 0:
        payable = 0

    payments = order.get('payments', []) or []
    paid_total = round(sum(p.get('amount', 0) for p in payments) + voucher_income_amount, 2)
    balance_due = round(payable - paid_total, 2)

    return {
        'subtotal': subtotal,
        'order_discount': order_discount,
        'order_discount_reason': order.get('order_discount_reason', ''),
        'discount_amount': discount_amount,
        'order_reduction': order_reduction,
        'order_reduction_reason': order.get('order_reduction_reason', ''),
        'voucher_name': voucher_name,
        'voucher_amount': voucher_amount,
        'voucher_income_amount': voucher_income_amount,
        'voucher_discount_amount': voucher_discount_amount,
        'voucher_items': voucher_items,
        'round_down': round_down,
        'payable': payable,
        'payments': payments,
        'paid_total': paid_total,
        'balance_due': balance_due,
    }


def attach_checkout_summary(order):
    """把 compute_checkout_summary 结果直接挂到 order 上并返回。"""
    order['checkout'] = compute_checkout_summary(order)
    return order


def financial_adjustments_locked(order):
    """已有收款且应收已结清时，必须先撤销部分收款才能调整金额。"""
    summary = compute_checkout_summary(order)
    return summary['paid_total'] > 0.005 and summary['balance_due'] <= 0.01


def financial_adjustment_exceeds_balance(order):
    """调整后的应收不得低于已收金额。"""
    summary = compute_checkout_summary(order)
    return summary['paid_total'] > summary['payable'] + 0.01


def financial_adjustment_locked_response():
    return jsonify({
        'success': False,
        'error': '订单已结清，请先撤销部分收款后再进行退菜或优惠操作',
    }), 400


def financial_adjustment_exceeded_response():
    return jsonify({
        'success': False,
        'error': '退菜或优惠金额不能超过当前还差金额',
    }), 400


def is_voucher_withdrawal(data):
    """撤销优惠券只会提高应收，结清后也允许单独执行。"""
    if set(data.keys()) != {'voucher'}:
        return False
    voucher = data.get('voucher') or {}
    if voucher.get('items') == []:
        return True
    return not voucher.get('name') and not (voucher.get('amount') or 0)


def current_time():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def append_operation_log(order, category, action, detail, operator='收银员'):
    """记录订单操作明细，供订单历史追溯。"""
    now = current_time()
    order.setdefault('operation_logs', []).append({
        'id': f"L{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
        'time': now,
        'category': category,
        'action': action,
        'detail': detail,
        'operator': operator,
    })


def build_order_export(orders):
    """生成包含订单、菜品和付款明细的 Excel。"""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError('Excel 导出组件尚未安装，请安装 openpyxl') from exc
    workbook = Workbook()
    order_sheet = workbook.active
    order_sheet.title = '订单汇总'
    order_sheet.append([
        '订单号', '桌台', '人数', '状态', '开台时间', '下单时间', '结束时间',
        '应收金额', '实收金额', '优惠金额', '付款方式', '订单备注',
    ])
    status_labels = {
        'pending': '待下单', 'submitted': '已下单', 'served': '已上菜',
        'paid': '已支付', 'canceled': '已取消', 'merged': '已并台',
    }
    item_sheet = workbook.create_sheet('菜品明细')
    item_sheet.append(['订单号', '桌台', '菜品', '数量', '单价', '小计', '下单时间', '标签', '备注'])
    payment_sheet = workbook.create_sheet('付款明细')
    payment_sheet.append(['订单号', '桌台', '类型', '付款方式', '金额', '时间', '备注'])

    for order in orders:
        summary = compute_checkout_summary(order)
        original = sum(
            (item.get('price', 0) or 0) * (item.get('quantity', 0) or 0)
            for item in order.get('items', []) if not item.get('returned')
        )
        saving = max(0, round(original - summary['payable'], 2))
        end_time = order.get('cleared_at') or order.get('canceled_at') or order.get('merged_at') or ''
        order_sheet.append([
            order.get('id'), order.get('table_id'), order.get('guests', 0),
            status_labels.get(order.get('status'), order.get('status')),
            order.get('created_at', ''), order.get('submitted_at', ''), end_time,
            summary['payable'], max(0, summary['paid_total'] - (order.get('refunded_amount', 0) or 0)),
            saving, order.get('payment_method', ''), order.get('remark', ''),
        ])
        for item in order.get('items', []):
            tags = []
            if item.get('gift_quantity'):
                tags.append(f"赠{item.get('gift_quantity')}")
            if item.get('discount'):
                tags.append(f"{(100 - item.get('discount')) / 10:g}折")
            if item.get('reduction'):
                tags.append(f"减¥{item.get('reduction')}")
            if item.get('returned'):
                tags.append(f"退菜：{item.get('return_reason', '')}")
            item_sheet.append([
                order.get('id'), order.get('table_id'), item.get('name'), item.get('quantity', 0),
                item.get('price', 0), item_line_total(item), item.get('added_at') or order.get('submitted_at', ''),
                '、'.join(tags), item.get('remark', ''),
            ])
        for payment in order.get('payments', []) or []:
            payment_sheet.append([
                order.get('id'), order.get('table_id'), '收款', payment.get('method'),
                payment.get('amount', 0), payment.get('time', ''), '',
            ])
        voucher_income = summary.get('voucher_income_amount', 0) or 0
        if voucher_income > 0:
            payment_sheet.append([
                order.get('id'), order.get('table_id'), '收款', '优惠券', voucher_income,
                order.get('paid_at') or order.get('submitted_at', ''),
                f"面值¥{summary.get('voucher_amount', 0):.2f}，优惠¥{summary.get('voucher_discount_amount', 0):.2f}",
            ])
        for refund in order.get('refunds', []) or []:
            payment_sheet.append([
                order.get('id'), order.get('table_id'), '退款', '原路退款', -abs(refund.get('amount', 0) or 0),
                refund.get('time', ''), refund.get('reason', ''),
            ])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(42, max(10, max_length + 2))
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_ticket_export(tickets):
    """生成制作单汇总与菜品明细 Excel。"""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError('Excel 导出组件尚未安装，请安装 openpyxl') from exc
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = '制作单汇总'
    summary_sheet.append([
        '制作单号', '订单号', '桌台', '类型', '状态', '菜品总数', '已完成数',
        '下单时间', '完成时间', '制作时长', '订单备注',
    ])
    item_sheet = workbook.create_sheet('菜品明细')
    item_sheet.append([
        '制作单号', '订单号', '桌台', '菜品', '数量', '菜品备注', '状态', '完成时间',
    ])
    for ticket in tickets:
        items = ticket.get('items', []) or []
        completed = [item for item in items if item.get('completed')]
        started_at = ticket.get('created_at') or ''
        ended_at = ticket.get('archived_at') or ''
        duration = ''
        try:
            start_time = datetime.strptime(started_at, '%Y-%m-%d %H:%M:%S')
            end_time = datetime.strptime(ended_at, '%Y-%m-%d %H:%M:%S') if ended_at else datetime.now()
            total_seconds = max(0, int((end_time - start_time).total_seconds()))
            duration = f'{total_seconds // 60:02d}:{total_seconds % 60:02d}'
        except (TypeError, ValueError):
            pass
        summary_sheet.append([
            ticket.get('id'), ticket.get('order_id'), ticket.get('table_id'), ticket.get('type'),
            '已完成' if ticket.get('archived') else '制作中',
            sum(item.get('quantity', 0) or 0 for item in items),
            sum(item.get('quantity', 0) or 0 for item in completed),
            started_at, ended_at, duration, ticket.get('order_remark', ''),
        ])
        for item in items:
            item_sheet.append([
                ticket.get('id'), ticket.get('order_id'), ticket.get('table_id'), item.get('name'),
                item.get('quantity', 0), item.get('remark', ''),
                '已完成' if item.get('completed') else '待出品', item.get('completed_at', ''),
            ])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(42, max(10, max_length + 2))
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def create_ticket(order, items=None, ticket_type='下单'):
    ticket_items = items if items is not None else order.get('items', [])
    if not ticket_items:
        return None

    now = current_time()
    ticket = {
        'id': f"T{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
        'type': ticket_type,
        'created_at': now,
        'archived': False,
        'order_remark': order.get('remark', ''),
        'items': [
            {
                'id': item.get('id'),
                'name': item.get('name'),
                'price': item.get('price'),
                'quantity': item.get('quantity', 0),
                'remark': item.get('remark', ''),
                'completed': False,
            }
            for item in ticket_items
            if item.get('quantity', 0) > 0
        ]
    }
    order.setdefault('tickets', []).append(ticket)
    return ticket

def generate_order_id(table_id):
    """生成订单编号: 日期-桌号-序号"""
    today = datetime.now().strftime('%Y%m%d')
    orders = load_orders()
    table_orders = [o for o in orders if o['table_id'] == table_id and o['id'].startswith(today)]
    seq = len(table_orders) + 1
    return f"{today}-{table_id}-{seq:03d}"

@order_bp.route('/<table_id>', methods=['GET'])
def get_order(table_id):
    """获取桌台当前订单"""
    orders = load_orders()
    order = find_active_order(orders, table_id)
    if not order:
        tables = load_tables()
        table = next((t for t in tables if t['id'] == table_id), None)
        if table and table.get('status') == 'pending_cleanup' and table.get('order_id'):
            order = next((o for o in orders if o['id'] == table.get('order_id')), None)
    if order:
        return jsonify({'success': True, 'data': attach_checkout_summary(order)})
    return jsonify({'success': True, 'data': None})

@order_bp.route('/<table_id>/create', methods=['POST'])
def create_order(table_id):
    """创建新订单"""
    orders = load_orders()

    # 检查是否已有未支付订单
    existing = find_active_order(orders, table_id)
    if existing:
        return jsonify({'success': True, 'data': existing})

    order = create_pending_order(table_id, orders)
    append_operation_log(order, '订单', '创建订单', f'桌台 {table_id}，人数 {order.get("guests", 0)}')
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})

@order_bp.route('/<table_id>/item', methods=['POST'])
def add_item(table_id):
    """添加菜品到订单"""
    data = request.json
    item_id = data.get('item_id')
    name = data.get('name')
    price = data.get('price')
    quantity = data.get('quantity', 1)
    add_mode_change = data.get('add_mode_change') is True

    if not item_id or not name or price is None:
        return jsonify({'success': False, 'error': '菜品参数不完整'}), 400

    try:
        price = float(price)
        quantity = int(quantity)
    except:
        return jsonify({'success': False, 'error': '菜品价格或数量无效'}), 400

    if quantity <= 0:
        return jsonify({'success': False, 'error': '菜品数量必须大于 0'}), 400

    orders = load_orders()
    active_order = find_active_order(orders, table_id)
    order = active_order

    if not order:
        order = create_pending_order(table_id, orders)
    elif 'guests' not in order:
        table = next((t for t in load_tables() if t['id'] == table_id), None)
        order['guests'] = table.get('guests', 0) if table else 0

    if order.get('status') != 'pending' and not add_mode_change:
        return jsonify({'success': False, 'error': '已下单菜品请通过加菜模式添加'}), 400

    # 加菜只合并本轮尚未提交的同款菜；绝不合并到上一轮已下单记录。
    existing_item = next((
        i for i in order['items']
        if (i.get('menu_item_id', i.get('id')) == item_id)
        and (not add_mode_change or i.get('addition_pending') is True)
    ), None)
    if existing_item:
        existing_item['quantity'] += quantity
    else:
        line_item = {
            'id': next_order_line_id(order) if add_mode_change else item_id,
            'menu_item_id': item_id,
            'name': name,
            'price': price,
            'quantity': quantity,
            'added_at': current_time(),
        }
        if add_mode_change:
            line_item['addition_pending'] = True
        # 新入列的菜品始终显示在明细顶部。
        order['items'].insert(0, line_item)

    # 重新计算总价
    recalc_order(order)
    append_operation_log(
        order,
        '菜品',
        '加菜' if add_mode_change else '添加菜品',
        f'{name} × {quantity}，单价 ¥{price:.2f}',
    )
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})

@order_bp.route('/<table_id>/item/<int:item_id>', methods=['DELETE'])
def remove_item(table_id, item_id):
    """从订单移除菜品"""
    data = request.get_json(silent=True) or {}
    orders = load_orders()
    active_order = find_active_order(orders, table_id)
    order = active_order

    if not order:
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    bypass_adjustment = data.get('add_mode_change') is True or data.get('add_mode_revert') is True
    if order.get('status') != 'pending' and not bypass_adjustment and financial_adjustments_locked(order):
        return financial_adjustment_locked_response()
    if order.get('status') != 'pending' and not data.get('reason') and not bypass_adjustment:
        return jsonify({'success': False, 'error': '请选择退菜原因'}), 400

    item = next((i for i in order['items'] if i['id'] == item_id), None)
    if not item:
        return jsonify({'success': False, 'error': '菜品不存在'}), 404

    soft_return = (
        order.get('status') != 'pending'
        and not bypass_adjustment
        and data.get('return_item') is True
    )
    if soft_return:
        item['returned'] = True
        item['return_reason'] = (data.get('reason') or '').strip() or '其他'
        item['returned_at'] = current_time()
    else:
        order['items'] = [i for i in order['items'] if i['id'] != item_id]
    recalc_order(order)
    if order.get('status') != 'pending' and not bypass_adjustment and financial_adjustment_exceeds_balance(order):
        return financial_adjustment_exceeded_response()
    append_operation_log(
        order,
        '菜品',
        '退菜' if soft_return else '删除菜品',
        f'{item.get("name", "菜品")} × {item.get("quantity", 0)}' + (f'，原因：{data.get("reason")}' if data.get('reason') else ''),
    )
    if order.get('status') != 'pending' and not bypass_adjustment:
        order.setdefault('adjustments', []).append({
            'type': 'remove',
            'item_id': item_id,
            'reason': data.get('reason'),
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})


@order_bp.route('/<table_id>/additions', methods=['DELETE'])
def clear_pending_additions(table_id):
    """取消或清空本轮尚未提交的加菜。"""
    orders = load_orders()
    order = find_active_order(orders, table_id)
    if not order:
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    pending_items = [item for item in order.get('items', []) if item.get('addition_pending') is True]
    order['items'] = [
        item for item in order.get('items', [])
        if item.get('addition_pending') is not True
    ]
    recalc_order(order)
    if pending_items:
        append_operation_log(
            order,
            '菜品',
            '取消本轮加菜',
            '、'.join(f'{item.get("name")} × {item.get("quantity", 0)}' for item in pending_items),
        )
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})

@order_bp.route('/<table_id>/item/<int:item_id>/quantity', methods=['PUT'])
def update_quantity(table_id, item_id):
    """更新菜品数量"""
    data = request.json or {}
    quantity = data.get('quantity', 1)
    try:
        quantity = int(quantity)
    except:
        return jsonify({'success': False, 'error': '菜品数量无效'}), 400

    orders = load_orders()
    active_order = find_active_order(orders, table_id)
    order = active_order

    if not order:
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    bypass_adjustment = data.get('add_mode_change') is True or data.get('add_mode_revert') is True
    if order.get('status') != 'pending' and not data.get('reason') and not bypass_adjustment:
        return jsonify({'success': False, 'error': '请选择改菜原因'}), 400

    item = next((i for i in order['items'] if i['id'] == item_id), None)
    if item and order.get('status') != 'pending' and not bypass_adjustment:
        if quantity < item.get('quantity', 0) and financial_adjustments_locked(order):
            return financial_adjustment_locked_response()
    if item:
        old_quantity = item.get('quantity', 0)
        if quantity <= 0:
            order['items'] = [i for i in order['items'] if i['id'] != item_id]
        else:
            item['quantity'] = quantity
            gift_qty = item.get('gift_quantity', 0) or 0
            if gift_qty > quantity:
                item['gift_quantity'] = quantity
        if order.get('status') != 'pending' and not bypass_adjustment:
            order.setdefault('adjustments', []).append({
                'type': 'quantity',
                'item_id': item_id,
                'from': old_quantity,
                'to': quantity,
                'reason': data.get('reason'),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })

    recalc_order(order)
    if order.get('status') != 'pending' and not bypass_adjustment and financial_adjustment_exceeds_balance(order):
        return financial_adjustment_exceeded_response()
    if item:
        append_operation_log(order, '菜品', '修改数量', f'{item.get("name", "菜品")}：{old_quantity} → {quantity}')
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})

@order_bp.route('/<table_id>/item/<int:item_id>', methods=['PATCH'])
def patch_item(table_id, item_id):
    """修改菜品属性：备注 / 折扣 / 减免"""
    data = request.get_json(silent=True) or {}
    orders = load_orders()
    order = find_active_order(orders, table_id)
    if not order:
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    item = next((i for i in order['items'] if i['id'] == item_id), None)
    if not item:
        return jsonify({'success': False, 'error': '菜品不存在'}), 404

    financial_fields = {'discount', 'reduction', 'gift_quantity'}
    has_financial_change = any(field in data for field in financial_fields)
    if has_financial_change and financial_adjustments_locked(order):
        return financial_adjustment_locked_response()

    if 'remark' in data:
        remark = (data.get('remark') or '').strip()
        if remark:
            item['remark'] = remark
        else:
            item.pop('remark', None)

    if 'discount' in data:
        try:
            discount = float(data.get('discount') or 0)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '折扣无效'}), 400
        if discount < 0 or discount > 100:
            return jsonify({'success': False, 'error': '折扣应在 0-100 之间'}), 400
        if discount == 0:
            item.pop('discount', None)
            item.pop('discount_reason', None)
        else:
            item['discount'] = discount
            discount_reason = (data.get('discount_reason') or '').strip()
            if discount_reason:
                item['discount_reason'] = discount_reason
            else:
                item.pop('discount_reason', None)

    if 'reduction' in data:
        try:
            reduction = float(data.get('reduction') or 0)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '减免金额无效'}), 400
        if reduction < 0:
            return jsonify({'success': False, 'error': '减免金额不能为负'}), 400
        line_base = item['price'] * item.get('quantity', 0)
        if reduction > line_base:
            return jsonify({'success': False, 'error': '减免金额不能超过菜品总额'}), 400
        if reduction == 0:
            item.pop('reduction', None)
            item.pop('reduction_reason', None)
        else:
            item['reduction'] = reduction
            reduction_reason = (data.get('reduction_reason') or '').strip()
            if reduction_reason:
                item['reduction_reason'] = reduction_reason
            else:
                item.pop('reduction_reason', None)

    if 'gift_quantity' in data:
        try:
            gift_qty = int(data.get('gift_quantity') or 0)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '赠菜数量无效'}), 400
        if gift_qty < 0:
            return jsonify({'success': False, 'error': '赠菜数量不能为负'}), 400
        if gift_qty > item.get('quantity', 0):
            return jsonify({'success': False, 'error': '赠菜数量不能超过菜品数量'}), 400
        if gift_qty == 0:
            item.pop('gift_quantity', None)
            item.pop('gift_reason', None)
        else:
            item['gift_quantity'] = gift_qty
            gift_reason = (data.get('gift_reason') or '').strip()
            if gift_reason:
                item['gift_reason'] = gift_reason
            else:
                item.pop('gift_reason', None)

    recalc_order(order)
    if has_financial_change and financial_adjustment_exceeds_balance(order):
        return financial_adjustment_exceeded_response()
    if 'remark' in data:
        append_operation_log(order, '菜品', '修改菜品备注', f'{item.get("name")}：{item.get("remark", "已清空") or "已清空"}')
    if 'discount' in data:
        append_operation_log(order, '优惠', '菜品折扣', f'{item.get("name")}：{(100 - (item.get("discount", 0) or 0)) / 10:g}折，原因：{item.get("discount_reason", "未填写")}')
    if 'reduction' in data:
        append_operation_log(order, '优惠', '菜品减免', f'{item.get("name")}：¥{item.get("reduction", 0):.2f}，原因：{item.get("reduction_reason", "未填写")}')
    if 'gift_quantity' in data:
        append_operation_log(order, '优惠', '赠菜', f'{item.get("name")}：{item.get("gift_quantity", 0)} 份，原因：{item.get("gift_reason", "未填写")}')
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})


@order_bp.route('/<table_id>/remark', methods=['PUT'])
def update_order_remark(table_id):
    """修改整单备注"""
    data = request.get_json(silent=True) or {}
    remark = (data.get('remark') or '').strip()
    orders = load_orders()
    order = find_active_order(orders, table_id)
    if not order:
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    if remark:
        order['remark'] = remark
    else:
        order.pop('remark', None)
    append_operation_log(order, '订单', '修改整单备注', remark or '清空整单备注')
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})

@order_bp.route('/<table_id>/submit', methods=['POST'])
def submit_order(table_id):
    """下单并生成虚拟小票，进入待结账"""
    orders = load_orders()
    order = next((o for o in orders if o['table_id'] == table_id and o['status'] == 'pending'), None)

    if not order:
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    if not order['items']:
        return jsonify({'success': False, 'error': '订单没有菜品，不能下单'}), 400

    order['status'] = 'submitted'
    order['submitted_at'] = current_time()
    append_operation_log(order, '订单', '提交下单', f'提交 {sum(item.get("quantity", 0) for item in order.get("items", []))} 份菜品')
    if production_ticket_enabled():
        create_ticket(order, ticket_type='下单')
    save_orders(orders)
    update_table_status(table_id, 'occupied', order_id=order['id'])
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})

@order_bp.route('/<table_id>/ticket', methods=['POST'])
def print_ticket(table_id):
    """补打或加菜打印虚拟小票"""
    data = request.json or {}
    orders = load_orders()
    order = find_active_order(orders, table_id)

    if not order:
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    if not order.get('items'):
        return jsonify({'success': False, 'error': '订单没有菜品，不能打印小票'}), 400

    ticket_type = data.get('type') or '补打'
    items = data.get('items')
    if items is not None and not isinstance(items, list):
        return jsonify({'success': False, 'error': '小票菜品格式错误'}), 400

    if not production_ticket_enabled():
        if ticket_type == '加菜':
            added = [item for item in order.get('items', []) if item.get('addition_pending') is True]
            for item in order.get('items', []):
                item.pop('addition_pending', None)
            append_operation_log(order, '订单', '确认加菜', f'确认 {sum(item.get("quantity", 0) for item in added)} 份加菜')
            save_orders(orders)
            return jsonify({
                'success': True,
                'data': {'order': attach_checkout_summary(order), 'ticket': None},
            })
        return jsonify({'success': False, 'error': '制作单模式未开启'}), 403

    ticket = create_ticket(order, items=items, ticket_type=ticket_type)
    if not ticket:
        return jsonify({'success': False, 'error': '小票没有菜品'}), 400

    if ticket_type == '加菜':
        added = [item for item in order.get('items', []) if item.get('addition_pending') is True]
        for item in order.get('items', []):
            item.pop('addition_pending', None)
        append_operation_log(order, '订单', '确认加菜', f'确认 {sum(item.get("quantity", 0) for item in added)} 份加菜')
    else:
        append_operation_log(order, '制作单', ticket_type, f'生成制作单 {ticket.get("id")}')
    save_orders(orders)
    return jsonify({'success': True, 'data': {'order': attach_checkout_summary(order), 'ticket': ticket}})

@order_bp.route('/<table_id>/ticket/<ticket_id>/archive', methods=['POST'])
def archive_ticket(table_id, ticket_id):
    """归档虚拟小票"""
    orders = load_orders()
    order = find_active_order(orders, table_id)

    if not order:
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    ticket = next((t for t in order.get('tickets', []) if t.get('id') == ticket_id), None)
    if not ticket:
        return jsonify({'success': False, 'error': '小票不存在'}), 404

    for item in ticket.get('items', []):
        item['completed'] = True
        item.setdefault('completed_at', current_time())
    ticket['archived'] = True
    ticket['archived_at'] = current_time()
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})


@order_bp.route('/<table_id>/ticket/<ticket_id>/items', methods=['PATCH'])
def update_ticket_items(table_id, ticket_id):
    """更新制作单划菜状态。"""
    data = request.get_json(silent=True) or {}
    orders = load_orders()
    order = next((o for o in orders if o.get('table_id') == table_id and any(
        ticket.get('id') == ticket_id for ticket in o.get('tickets', [])
    )), None)
    if not order:
        return jsonify({'success': False, 'error': '制作单不存在'}), 404
    ticket = next((t for t in order.get('tickets', []) if t.get('id') == ticket_id), None)
    items = ticket.get('items', [])
    completed = data.get('completed') is True
    if data.get('all') is True:
        targets = items
    else:
        try:
            item_index = int(data.get('item_index'))
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '制作单菜品序号无效'}), 400
        if item_index < 0 or item_index >= len(items):
            return jsonify({'success': False, 'error': '制作单菜品不存在'}), 404
        targets = [items[item_index]]
    for item in targets:
        item['completed'] = completed
        if completed:
            item['completed_at'] = current_time()
        else:
            item.pop('completed_at', None)
    save_orders(orders)
    return jsonify({'success': True, 'data': ticket})

@order_bp.route('/tickets', methods=['GET'])
def list_tickets():
    """获取未归档虚拟小票"""
    if not production_ticket_enabled():
        return jsonify({'success': True, 'data': []})
    orders = load_orders()
    tickets = []
    for order in orders:
        if order.get('status') not in ['submitted', 'served']:
            continue
        for ticket in order.get('tickets', []):
            if ticket.get('archived'):
                continue
            tickets.append({
                'id': ticket.get('id'),
                'type': ticket.get('type', '下单'),
                'created_at': ticket.get('created_at'),
                'table_id': order.get('table_id'),
                'guests': order.get('guests', 0),
                'order_remark': ticket.get('order_remark') or order.get('remark', ''),
                'items': ticket.get('items', [])
            })

    tickets.sort(key=lambda item: item.get('created_at') or '', reverse=True)
    return jsonify({'success': True, 'data': tickets})


@order_bp.route('/tickets/history', methods=['GET'])
def ticket_history():
    """获取全部历史制作单，包括进行中和已完成。"""
    if not production_ticket_enabled():
        return jsonify({'success': True, 'data': []})
    orders = load_orders()
    tickets = []
    for order in orders:
        for ticket in order.get('tickets', []):
            tickets.append({
                'id': ticket.get('id'),
                'order_id': order.get('id'),
                'type': ticket.get('type', '下单'),
                'created_at': ticket.get('created_at'),
                'archived': ticket.get('archived', False),
                'archived_at': ticket.get('archived_at'),
                'table_id': order.get('table_id'),
                'guests': order.get('guests', 0),
                'order_remark': ticket.get('order_remark') or order.get('remark', ''),
                'items': ticket.get('items', []),
            })
    tickets.sort(key=lambda item: item.get('created_at') or '', reverse=True)
    return jsonify({'success': True, 'data': tickets})


@order_bp.route('/tickets/export', methods=['POST'])
def export_ticket_history():
    data = request.get_json(silent=True) or {}
    requested_ids = data.get('ticket_ids')
    if not isinstance(requested_ids, list):
        return jsonify({'success': False, 'error': '制作单参数无效'}), 400
    selected_ids = set(str(ticket_id) for ticket_id in requested_ids)
    tickets = []
    for order in load_orders():
        for ticket in order.get('tickets', []):
            if str(ticket.get('id')) not in selected_ids:
                continue
            tickets.append({
                **ticket,
                'order_id': order.get('id'),
                'table_id': order.get('table_id'),
                'order_remark': ticket.get('order_remark') or order.get('remark', ''),
            })
    if not tickets:
        return jsonify({'success': False, 'error': '没有可导出的制作单'}), 400
    tickets.sort(key=lambda item: item.get('created_at') or '', reverse=True)
    try:
        output = build_ticket_export(tickets)
    except RuntimeError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 503
    filename = f"制作单记录-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

@order_bp.route('/<table_id>/checkout-config', methods=['PATCH'])
def update_checkout_config(table_id):
    """设置整单折扣、整单减免、抵扣券、抹零"""
    orders = load_orders()
    order = next((o for o in orders if o['table_id'] == table_id and o.get('status') in ['submitted', 'served']), None)
    if not order:
        return jsonify({'success': False, 'error': '订单不在可结账状态'}), 404

    data = request.get_json(silent=True) or {}
    if data and financial_adjustments_locked(order) and not is_voucher_withdrawal(data):
        return financial_adjustment_locked_response()

    if 'order_discount' in data:
        try:
            v = float(data.get('order_discount') or 0)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '整单折扣无效'}), 400
        if v < 0 or v > 100:
            return jsonify({'success': False, 'error': '整单折扣应在 0-100 之间'}), 400
        if v == 0:
            order.pop('order_discount', None)
            order.pop('order_discount_reason', None)
        else:
            order['order_discount'] = v
            reason = (data.get('order_discount_reason') or '').strip()
            if reason:
                order['order_discount_reason'] = reason
            else:
                order.pop('order_discount_reason', None)

    if 'order_reduction' in data:
        try:
            v = round(float(data.get('order_reduction') or 0), 2)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '整单减免无效'}), 400
        if v < 0:
            return jsonify({'success': False, 'error': '整单减免不能为负'}), 400
        subtotal = round(order.get('total', 0) or 0, 2)
        discount = order.get('order_discount', 0) or 0
        max_reduction = round(subtotal * (1 - discount / 100), 2)
        if v > max_reduction:
            return jsonify({'success': False, 'error': '整单减免不能超过折后订单金额'}), 400
        if v == 0:
            order.pop('order_reduction', None)
            order.pop('order_reduction_reason', None)
        else:
            order['order_reduction'] = v
            reason = (data.get('order_reduction_reason') or '').strip()
            if reason:
                order['order_reduction_reason'] = reason
            else:
                order.pop('order_reduction_reason', None)

    if 'voucher' in data:
        voucher_data = data.get('voucher') or {}
        requested_items = voucher_data.get('items')
        if isinstance(requested_items, list):
            definitions = {v.get('id'): v for v in load_vouchers()}
            voucher_items = []
            for requested in requested_items:
                try:
                    voucher_id = int(requested.get('id'))
                    quantity = int(requested.get('quantity') or 0)
                except (AttributeError, TypeError, ValueError):
                    return jsonify({'success': False, 'error': '优惠券使用数量无效'}), 400
                if quantity <= 0:
                    continue
                if quantity > 99:
                    return jsonify({'success': False, 'error': '单种优惠券最多使用 99 张'}), 400
                definition = definitions.get(voucher_id)
                if not definition:
                    return jsonify({'success': False, 'error': '优惠券不存在或已删除'}), 400
                face_value = round(float(definition.get('face_value') or 0), 2)
                sale_price = round(float(definition.get('sale_price') or face_value), 2)
                voucher_items.append({
                    'id': voucher_id,
                    'name': definition.get('name') or '',
                    'quantity': quantity,
                    'face_value': face_value,
                    'sale_price': sale_price,
                    'amount': round(face_value * quantity, 2),
                })
            voucher_amount = round(sum(item['amount'] for item in voucher_items), 2)
            if voucher_amount <= 0:
                order.pop('voucher', None)
            else:
                voucher_name = '、'.join(
                    f"{item['name']}×{item['quantity']}" for item in voucher_items
                )
                order['voucher'] = {
                    'name': voucher_name,
                    'amount': voucher_amount,
                    'items': voucher_items,
                }
        else:
            voucher_name = (voucher_data.get('name') or '').strip()
            try:
                voucher_amount = round(float(voucher_data.get('amount') or 0), 2)
            except (TypeError, ValueError):
                return jsonify({'success': False, 'error': '抵扣券金额无效'}), 400
            if voucher_amount < 0:
                return jsonify({'success': False, 'error': '抵扣券金额不能为负'}), 400
            if voucher_amount == 0 and not voucher_name:
                order.pop('voucher', None)
            else:
                order['voucher'] = {'name': voucher_name, 'amount': voucher_amount}

    if 'round_down' in data:
        try:
            v = round(float(data.get('round_down') or 0), 2)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '抹零金额无效'}), 400
        if v < 0:
            return jsonify({'success': False, 'error': '抹零金额不能为负'}), 400
        if v == 0:
            order.pop('round_down', None)
        else:
            order['round_down'] = v

    if data.get('clear_all_offers'):
        for item in order.get('items', []):
            for field in (
                'discount', 'discount_reason',
                'reduction', 'reduction_reason',
                'gift_quantity', 'gift_reason',
            ):
                item.pop(field, None)
        recalc_order(order)
        order.pop('order_reduction', None)
        order.pop('order_reduction_reason', None)
        order.pop('voucher', None)
        order.pop('round_down', None)

    if financial_adjustment_exceeds_balance(order):
        return financial_adjustment_exceeded_response()

    if 'order_discount' in data:
        discount = order.get('order_discount', 0) or 0
        append_operation_log(order, '优惠', '整单折扣', f'{(100 - discount) / 10:g}折，原因：{order.get("order_discount_reason", "未填写")}')
    if 'order_reduction' in data:
        append_operation_log(order, '优惠', '整单减免', f'¥{order.get("order_reduction", 0):.2f}，原因：{order.get("order_reduction_reason", "未填写")}')
    if 'voucher' in data:
        voucher = order.get('voucher') or {}
        voucher_face, voucher_income, voucher_discount = voucher_financials(voucher)
        append_operation_log(
            order,
            '优惠',
            '使用优惠券' if voucher else '撤销优惠券',
            f'{voucher.get("name")}，面值 ¥{voucher_face:.2f}，实收 ¥{voucher_income:.2f}，优惠 ¥{voucher_discount:.2f}' if voucher else '恢复优惠券抵扣金额',
        )
    if 'round_down' in data:
        append_operation_log(order, '优惠', '抹零', f'¥{order.get("round_down", 0):.2f}')
    if data.get('clear_all_offers'):
        append_operation_log(order, '优惠', '清除原优惠', '免单前清除菜品及整单原有优惠')

    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})


@order_bp.route('/<table_id>/payment', methods=['POST'])
def add_payment(table_id):
    """新增一笔收款记录（支持组合支付）"""
    orders = load_orders()
    order = next((o for o in orders if o['table_id'] == table_id and o.get('status') in ['submitted', 'served']), None)
    if not order:
        return jsonify({'success': False, 'error': '订单不在可结账状态'}), 404

    data = request.get_json(silent=True) or {}
    method = (data.get('method') or '').strip()
    if not method:
        return jsonify({'success': False, 'error': '请选择付款方式'}), 400
    if method not in ALLOWED_PAYMENT_METHODS:
        return jsonify({'success': False, 'error': '付款方式无效'}), 400

    try:
        amount = round(float(data.get('amount') or 0), 2)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '收款金额无效'}), 400
    if amount <= 0:
        return jsonify({'success': False, 'error': '收款金额必须大于 0'}), 400

    payments = order.setdefault('payments', [])
    payment = {
        'id': f"P{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
        'method': method,
        'amount': amount,
        'time': current_time(),
    }
    payments.append(payment)
    append_operation_log(order, '收款', '新增收款', f'{method} ¥{amount:.2f}，收款流水 {payment["id"]}')
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})


@order_bp.route('/<table_id>/payment/<payment_id>', methods=['DELETE'])
def revert_payment(table_id, payment_id):
    """撤销一笔已收款"""
    orders = load_orders()
    order = next((o for o in orders if o['table_id'] == table_id and o.get('status') in ['submitted', 'served']), None)
    if not order:
        return jsonify({'success': False, 'error': '订单不在可结账状态'}), 404

    payments = order.get('payments', []) or []
    target = next((p for p in payments if p.get('id') == payment_id), None)
    if not target:
        return jsonify({'success': False, 'error': '收款记录不存在'}), 404

    order['payments'] = [p for p in payments if p.get('id') != payment_id]
    if not order['payments']:
        order.pop('payments', None)
    append_operation_log(order, '收款', '撤销收款', f'{target.get("method")} ¥{target.get("amount", 0):.2f}，收款流水 {payment_id}')
    save_orders(orders)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})


@order_bp.route('/<table_id>/checkout', methods=['POST'])
def checkout(table_id):
    """正式结账清台：将已收款确认完毕的订单标记为已支付，桌台进入待清台"""
    orders = load_orders()
    order = next((o for o in orders if o['table_id'] == table_id and o['status'] in ['submitted', 'served']), None)

    if not order:
        return jsonify({'success': False, 'error': '订单不在可结账状态'}), 404

    if not order['items']:
        return jsonify({'success': False, 'error': '订单没有菜品，不能结账'}), 400

    if 'guests' not in order:
        table = next((t for t in load_tables() if t['id'] == table_id), None)
        order['guests'] = table.get('guests', 0) if table else 0

    summary = compute_checkout_summary(order)
    payable = summary['payable']
    payments = list(order.get('payments', []) or [])

    data = request.get_json(silent=True) or {}

    # 如果请求体带了 payment_method 且当前没有收款记录，则把这一笔作为最终收款一次性记入
    if not payments and data.get('payment_method'):
        method = (data.get('payment_method') or '').strip()
        if method not in ALLOWED_PAYMENT_METHODS:
            return jsonify({'success': False, 'error': '付款方式无效'}), 400
        try:
            amount = round(float(data.get('paid_amount') or summary['balance_due']), 2)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '实收金额无效'}), 400
        if amount <= 0:
            return jsonify({'success': False, 'error': '实收金额必须大于 0'}), 400
        payments.append({
            'id': f"P{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
            'method': method,
            'amount': amount,
            'time': current_time(),
        })
        order['payments'] = payments

    voucher_income = summary.get('voucher_income_amount', 0) or 0
    paid_total = round(sum(p.get('amount', 0) for p in payments) + voucher_income, 2)
    if paid_total < payable - 0.01:
        return jsonify({
            'success': False,
            'error': f'未付清：应收 ¥{payable}，已收 ¥{paid_total}，差额 ¥{round(payable - paid_total, 2)}'
        }), 400

    # 多收金额作为小费/溢收，不退还；最终以 payable 作为 paid_amount 记账
    methods = [p.get('method') for p in payments]
    if voucher_income > 0:
        methods.append('优惠券')
    methods = list(dict.fromkeys(methods))
    if len(methods) > 1:
        payment_method = '组合支付'
    elif methods:
        payment_method = methods[0]
    else:
        payment_method = (data.get('payment_method') or '').strip() or '其他'

    order['status'] = 'paid'
    order['paid_at'] = current_time()
    order['payment_method'] = payment_method
    order['paid_amount'] = paid_total
    order['payments'] = payments
    append_operation_log(order, '订单', '完成结账', f'应收 ¥{payable:.2f}，实收 ¥{paid_total:.2f}，{payment_method}')
    save_orders(orders)
    update_table_status(table_id, 'pending_cleanup', order_id=order['id'])
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})

@order_bp.route('/<table_id>/transfer', methods=['POST'])
def transfer_table(table_id):
    """转台：将当前桌台的订单整体迁移到另一空桌"""
    data = request.get_json(silent=True) or {}
    target_id = data.get('target_table_id')
    if not target_id:
        return jsonify({'success': False, 'error': '请选择目标桌台'}), 400
    if target_id == table_id:
        return jsonify({'success': False, 'error': '目标桌台不能与当前桌台相同'}), 400

    tables = load_tables()
    src = next((t for t in tables if t['id'] == table_id), None)
    dst = next((t for t in tables if t['id'] == target_id), None)
    if not src:
        return jsonify({'success': False, 'error': '当前桌台不存在'}), 404
    if not dst:
        return jsonify({'success': False, 'error': '目标桌台不存在'}), 404
    if dst.get('status') != 'empty':
        return jsonify({'success': False, 'error': '目标桌台必须为空桌'}), 400

    orders = load_orders()
    order = find_active_order(orders, table_id)
    if not order:
        return jsonify({'success': False, 'error': '当前桌台没有活动订单'}), 404

    order['table_id'] = target_id
    order['transferred_from'] = table_id
    order['transferred_at'] = current_time()
    append_operation_log(order, '桌台', '转台', f'{table_id} → {target_id}')

    src_guests = src.get('guests', 0)
    src_opened = src.get('opened_at')

    src['status'] = 'empty'
    src['guests'] = 0
    src['opened_at'] = None
    src['order_id'] = None

    dst['status'] = 'occupied'
    dst['guests'] = src_guests or order.get('guests', 1)
    dst['opened_at'] = src_opened or current_time()
    dst['order_id'] = order['id']

    save_orders(orders)
    save_tables(tables)
    return jsonify({'success': True, 'data': attach_checkout_summary(order)})


@order_bp.route('/<table_id>/merge', methods=['POST'])
def merge_table(table_id):
    """并台：将当前桌台的订单合并到另一已开台的桌台"""
    data = request.get_json(silent=True) or {}
    target_id = data.get('target_table_id')
    if not target_id:
        return jsonify({'success': False, 'error': '请选择目标桌台'}), 400
    if target_id == table_id:
        return jsonify({'success': False, 'error': '目标桌台不能与当前桌台相同'}), 400

    tables = load_tables()
    src = next((t for t in tables if t['id'] == table_id), None)
    dst = next((t for t in tables if t['id'] == target_id), None)
    if not src:
        return jsonify({'success': False, 'error': '当前桌台不存在'}), 404
    if not dst:
        return jsonify({'success': False, 'error': '目标桌台不存在'}), 404
    if dst.get('status') != 'occupied':
        return jsonify({'success': False, 'error': '目标桌台必须为已开台'}), 400

    orders = load_orders()
    src_order = find_active_order(orders, table_id)
    dst_order = find_active_order(orders, target_id)
    if not src_order:
        return jsonify({'success': False, 'error': '当前桌台没有活动订单'}), 404
    if not dst_order:
        return jsonify({'success': False, 'error': '目标桌台没有活动订单'}), 404

    for src_item in src_order.get('items', []):
        existing = next((i for i in dst_order['items'] if i['id'] == src_item['id']), None)
        if existing:
            existing['quantity'] += src_item.get('quantity', 0)
            existing.pop('discount', None)
            existing.pop('reduction', None)
            if src_item.get('remark'):
                existing['remark'] = src_item['remark']
            else:
                existing.pop('remark', None)
        else:
            new_item = {
                'id': src_item['id'],
                'name': src_item['name'],
                'price': src_item['price'],
                'quantity': src_item.get('quantity', 0),
                'added_at': src_item.get('added_at') or src_order.get('submitted_at') or src_order.get('created_at'),
            }
            if src_item.get('remark'):
                new_item['remark'] = src_item['remark']
            dst_order['items'].append(new_item)

    recalc_order(dst_order)

    merged_quantity = sum(item.get('quantity', 0) for item in src_order.get('items', []))
    append_operation_log(src_order, '桌台', '并台转出', f'{table_id} → {target_id}，转出 {merged_quantity} 份菜品')
    append_operation_log(dst_order, '桌台', '并台接收', f'接收桌台 {table_id} 的 {merged_quantity} 份菜品')

    src_order['status'] = 'merged'
    src_order['merged_to'] = dst_order['id']
    src_order['merged_at'] = current_time()
    src_order['items'] = []

    src['status'] = 'empty'
    src['guests'] = 0
    src['opened_at'] = None
    src['order_id'] = None

    save_orders(orders)
    save_tables(tables)
    return jsonify({'success': True, 'data': attach_checkout_summary(dst_order)})


@order_bp.route('/<table_id>/cancel', methods=['POST'])
def cancel_order(table_id):
    """取消当前订单并释放桌台"""
    orders = load_orders()
    order = find_active_order(orders, table_id)

    if order:
        withdrawing_table = order.get('status') == 'pending'
        order['status'] = 'canceled'
        order['canceled_at'] = current_time()
        append_operation_log(
            order,
            '桌台' if withdrawing_table else '订单',
            '撤台' if withdrawing_table else '撤销订单',
            f'桌台 {table_id} 已撤台并恢复为空桌' if withdrawing_table else f'桌台 {table_id} 订单已取消',
        )
        save_orders(orders)

    update_table_status(table_id, 'empty', reset=True)
    return jsonify({
        'success': True,
        'data': attach_checkout_summary(order) if order else None,
    })

@order_bp.route('/list', methods=['GET'])
def list_orders():
    """获取订单历史"""
    orders = load_orders()
    # 按时间倒序
    orders.sort(key=lambda x: x.get('paid_at') or x.get('created_at'), reverse=True)
    return jsonify({'success': True, 'data': [attach_checkout_summary(order) for order in orders]})


def selected_export_orders(order_ids):
    orders = load_orders()
    if not isinstance(order_ids, list):
        return []
    selected = set(str(order_id) for order_id in order_ids)
    return [order for order in orders if str(order.get('id')) in selected]


@order_bp.route('/export', methods=['POST'])
def export_orders():
    data = request.get_json(silent=True) or {}
    orders = selected_export_orders(data.get('order_ids'))
    if not orders:
        return jsonify({'success': False, 'error': '没有可导出的订单'}), 400
    try:
        output = build_order_export(orders)
    except RuntimeError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 503
    filename = f"订单历史-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@order_bp.route('/export/email', methods=['POST'])
def email_order_export():
    data = request.get_json(silent=True) or {}
    recipient = (data.get('email') or '').strip()
    if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', recipient):
        return jsonify({'success': False, 'error': '收件邮箱格式不正确'}), 400
    orders = selected_export_orders(data.get('order_ids'))
    if not orders:
        return jsonify({'success': False, 'error': '没有可导出的订单'}), 400

    smtp_host = os.environ.get('SMTP_HOST', '').strip()
    smtp_user = os.environ.get('SMTP_USER', '').strip()
    smtp_password = os.environ.get('SMTP_PASSWORD', '')
    smtp_from = os.environ.get('SMTP_FROM', '').strip() or smtp_user
    if not smtp_host or not smtp_from:
        return jsonify({
            'success': False,
            'error': '邮件导出尚未配置发件邮箱，请配置 SMTP_HOST、SMTP_USER、SMTP_PASSWORD 和 SMTP_FROM',
        }), 503
    try:
        smtp_port = int(os.environ.get('SMTP_PORT', '465'))
    except ValueError:
        return jsonify({'success': False, 'error': 'SMTP_PORT 配置无效'}), 500

    filename = f"订单历史-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    try:
        output = build_order_export(orders)
    except RuntimeError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 503
    message = EmailMessage()
    message['Subject'] = '订单历史导出'
    message['From'] = smtp_from
    message['To'] = recipient
    message.set_content(f'附件为订单历史导出，共 {len(orders)} 笔订单。')
    message.add_attachment(
        output.getvalue(),
        maintype='application',
        subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=filename,
    )
    try:
        use_ssl = os.environ.get('SMTP_SSL', 'true').lower() == 'true'
        if use_ssl:
            server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=20)
        else:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=20)
            if os.environ.get('SMTP_USE_TLS', 'true').lower() == 'true':
                server.starttls()
        with server:
            if smtp_user:
                server.login(smtp_user, smtp_password)
            server.send_message(message)
    except Exception:
        return jsonify({'success': False, 'error': '邮件发送失败，请检查发件邮箱和 SMTP 配置'}), 502
    return jsonify({'success': True, 'data': {'email': recipient, 'count': len(orders)}})


@order_bp.route('/history/<order_id>/refund', methods=['POST'])
def refund_history_order(order_id):
    """为已支付订单登记退款，支持多次部分退款。"""
    orders = load_orders()
    order = next((o for o in orders if o.get('id') == order_id), None)
    if not order:
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    if order.get('status') != 'paid':
        return jsonify({'success': False, 'error': '只有已支付订单可以退款'}), 400

    data = request.get_json(silent=True) or {}
    reason = (data.get('reason') or '').strip()
    if not reason:
        return jsonify({'success': False, 'error': '请填写退款原因'}), 400
    try:
        amount = round(float(data.get('amount') or 0), 2)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '退款金额无效'}), 400
    if amount <= 0:
        return jsonify({'success': False, 'error': '退款金额必须大于 0'}), 400

    payments = order.get('payments', []) or []
    voucher_income = voucher_financials(order.get('voucher'))[1]
    paid_total = round(sum(p.get('amount', 0) or 0 for p in payments) + voucher_income, 2)
    if not payments and voucher_income <= 0:
        paid_total = round(order.get('paid_amount', order.get('total', 0)) or 0, 2)
    refunds = order.setdefault('refunds', [])
    refunded_total = round(sum(r.get('amount', 0) or 0 for r in refunds), 2)
    refundable = round(max(0, paid_total - refunded_total), 2)
    if amount > refundable + 0.001:
        return jsonify({'success': False, 'error': f'最多可退款 ¥{refundable:.2f}'}), 400

    refunds.append({
        'id': f"R{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
        'amount': amount,
        'reason': reason,
        'time': current_time(),
    })
    total_after = round(refunded_total + amount, 2)
    order['refund_status'] = 'full' if total_after >= paid_total - 0.01 else 'partial'
    order['refunded_amount'] = total_after
    append_operation_log(order, '退款', '订单退款', f'退款 ¥{amount:.2f}，原因：{reason}')
    save_orders(orders)
    return jsonify({'success': True, 'data': order})

@order_bp.route('/stats', methods=['GET'])
def stats():
    """获取日结统计"""
    today = datetime.now().strftime('%Y%m%d')
    orders = load_orders()
    today_orders = [o for o in orders if o['status'] == 'paid' and o['id'].startswith(today)]
    pending_orders = [o for o in orders if o.get('status') in ['pending', 'submitted', 'served']]
    tables = load_tables()
    opened_tables = [t for t in tables if t['status'] == 'occupied']

    total_amount = round(sum(o['total'] for o in today_orders), 2)
    total_count = len(today_orders)
    total_guests = sum(o.get('guests', 0) for o in today_orders)
    unpaid_amount = round(sum(committed_order_total(o) for o in pending_orders), 2)
    unpaid_count = len(pending_orders)
    unpaid_guests = sum(t.get('guests', 0) for t in opened_tables)

    return jsonify({
        'success': True,
        'data': {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'total_amount': total_amount,
            'total_count': total_count,
            'total_guests': total_guests,
            'unpaid_amount': unpaid_amount,
            'unpaid_count': unpaid_count,
            'unpaid_guests': unpaid_guests
        }
    })
