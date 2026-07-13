from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO

from flask import Blueprint, g, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select

from auth.database import get_auth_session
from auth.middleware import require_permission
from auth.models import AuditLog, User
from auth.permissions import utc_iso
from auth.service import audit


audit_logs_bp = Blueprint('audit_logs', __name__)


def _parse_datetime(value: str):
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace('Z', '+00:00'))
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone().replace(tzinfo=None)
        return parsed
    except ValueError:
        return None


def _apply_filters(statement, values):
    if not any(role.code == 'superadmin' for role in g.current_user.roles):
        statement = statement.where(AuditLog.role_snapshot.not_like('%"superadmin"%'))
        db = get_auth_session()
        superadmin_ids = [
            user.id for user in db.scalars(select(User)).all()
            if any(role.code == 'superadmin' for role in user.roles)
        ]
        for user_id in superadmin_ids:
            pattern = f'%{user_id}%'
            statement = statement.where(
                or_(AuditLog.before_snapshot.is_(None), AuditLog.before_snapshot.not_like(pattern)),
                or_(AuditLog.after_snapshot.is_(None), AuditLog.after_snapshot.not_like(pattern)),
                or_(AuditLog.resource_id.is_(None), AuditLog.resource_id != user_id),
                or_(AuditLog.approver_user_id.is_(None), AuditLog.approver_user_id != user_id),
            )
    keyword = str(values.get('keyword', '')).strip()
    module = str(values.get('module', '')).strip()
    result = str(values.get('result', '')).strip()
    action = str(values.get('action', '')).strip()
    resource = str(values.get('resource', '')).strip()
    date_from = _parse_datetime(str(values.get('date_from', '')).strip())
    date_to = _parse_datetime(str(values.get('date_to', '')).strip())
    if keyword:
        like = f'%{keyword}%'
        statement = statement.where(or_(
            AuditLog.user_name.like(like),
            AuditLog.action.like(like),
            AuditLog.resource_id.like(like),
            AuditLog.reason.like(like),
        ))
    if module:
        statement = statement.where(AuditLog.module == module)
    if result:
        statement = statement.where(AuditLog.result == result)
    if action:
        statement = statement.where(AuditLog.action.like(f'%{action}%'))
    if resource:
        statement = statement.where(or_(AuditLog.resource_type.like(f'%{resource}%'), AuditLog.resource_id.like(f'%{resource}%')))
    if date_from:
        statement = statement.where(AuditLog.created_at >= date_from)
    if date_to:
        statement = statement.where(AuditLog.created_at <= date_to)
    return statement


def _json_value(value):
    if not value:
        return None
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return value


def _serialize(item: AuditLog) -> dict:
    return {
        'id': item.id,
        'created_at': utc_iso(item.created_at),
        'user_id': item.user_id,
        'user_name': item.user_name,
        'roles': _json_value(item.role_snapshot) or [],
        'ip_address': item.ip_address,
        'user_agent': item.user_agent,
        'module': item.module,
        'action': item.action,
        'resource_type': item.resource_type,
        'resource_id': item.resource_id,
        'before': _json_value(item.before_snapshot),
        'after': _json_value(item.after_snapshot),
        'reason': item.reason,
        'approver_user_id': item.approver_user_id,
        'result': item.result,
        'error_code': item.error_code,
    }


@audit_logs_bp.get('')
@require_permission('audit.view')
def list_audit_logs():
    db = get_auth_session()
    try:
        page = max(1, int(request.args.get('page', 1)))
        page_size = min(100, max(10, int(request.args.get('page_size', 20))))
    except ValueError:
        return jsonify({'success': False, 'error': '分页参数无效'}), 400
    filtered = _apply_filters(select(AuditLog), request.args)
    count_statement = _apply_filters(select(func.count(AuditLog.id)), request.args)
    total = int(db.scalar(count_statement) or 0)
    records = list(db.scalars(
        filtered.order_by(AuditLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    ).all())
    module_statement = select(AuditLog.module).distinct().order_by(AuditLog.module.asc())
    if not any(role.code == 'superadmin' for role in g.current_user.roles):
        module_statement = module_statement.where(AuditLog.role_snapshot.not_like('%"superadmin"%'))
    modules = list(db.scalars(module_statement).all())
    return jsonify({
        'success': True,
        'data': [_serialize(item) for item in records],
        'total': total,
        'page': page,
        'page_size': page_size,
        'modules': modules,
    })


@audit_logs_bp.post('/export')
@require_permission('audit.export')
def export_audit_logs():
    db = get_auth_session()
    values = request.get_json(silent=True) or {}
    records = list(db.scalars(
        _apply_filters(select(AuditLog), values).order_by(AuditLog.created_at.desc()).limit(10000)
    ).all())
    audit(
        db,
        request,
        action='audit.export',
        module='audit',
        user=g.current_user,
        resource_type='audit_log',
        after={'record_count': len(records), 'filters': values},
    )
    db.commit()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '系统日志'
    headers = ['序号', '时间', '操作人', '角色', '模块', '动作', '业务对象', '对象ID', 'IP', '结果', '错误码', '原因', '变更前', '变更后']
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4F46E5')
    for index, item in enumerate(records, 1):
        data = _serialize(item)
        worksheet.append([
            index,
            item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            item.user_name or '匿名/系统',
            ' / '.join(data['roles']),
            item.module,
            item.action,
            item.resource_type or '',
            item.resource_id or '',
            item.ip_address or '',
            item.result,
            item.error_code or '',
            item.reason or '',
            json.dumps(data['before'], ensure_ascii=False) if data['before'] is not None else '',
            json.dumps(data['after'], ensure_ascii=False) if data['after'] is not None else '',
        ])
    widths = [8, 22, 16, 22, 14, 24, 16, 22, 16, 12, 16, 24, 42, 42]
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    worksheet.freeze_panes = 'A2'
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f'silver-lining-system-logs-{datetime.now().strftime("%Y%m%d-%H%M%S")}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
